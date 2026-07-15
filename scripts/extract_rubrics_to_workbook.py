#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import sys
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass
from html import unescape
from html.parser import HTMLParser
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from common_xml import load_export_root, resolve_export_root


GREEN_FILL = PatternFill(fill_type="solid", fgColor="D9EAD3")
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")

RUBRICS_SCHEMA = "coursecraft.rubrics/1"


class HtmlToTextParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.parts: list[str] = []

    def handle_starttag(self, tag: str, attrs) -> None:  # type: ignore[override]
        if tag in {"p", "br", "div", "tr", "li", "td"}:
            self.parts.append("\n")

    def handle_endtag(self, tag: str) -> None:  # type: ignore[override]
        if tag in {"p", "div", "tr", "li", "td"}:
            self.parts.append("\n")

    def handle_data(self, data: str) -> None:  # type: ignore[override]
        self.parts.append(data)

    def get_text(self) -> str:
        raw = "".join(self.parts)
        lines = [" ".join(line.split()) for line in raw.splitlines()]
        lines = [line for line in lines if line]
        return "\n".join(lines).strip()


@dataclass
class Level:
    level_id: str
    name: str
    sort_order: int
    score_band: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract one or more Brightspace rubrics from rubrics_d2l.xml into a workbook shaped like the Points and Rubrics download."
    )
    parser.add_argument(
        "source",
        help="Path to rubrics_d2l.xml or to a folder containing rubrics_d2l.xml",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Optional output .xlsx path. Defaults to <folder>/Rubrics Extract.xlsx",
    )
    parser.add_argument(
        "--one-sheet-per-rubric",
        action="store_true",
        help="Write each rubric to its own worksheet instead of stacking all rubrics on one Rubrics sheet.",
    )
    parser.add_argument(
        "--json",
        action="store_true",
        help=f"Also write the rubric grids as canonical JSON ({RUBRICS_SCHEMA}).",
    )
    parser.add_argument(
        "--json-output",
        default=None,
        help="Optional output .json path (implies --json). Defaults to the workbook path with a .json suffix.",
    )
    return parser.parse_args()


def locate_rubrics_xml(source: Path, temp_dirs: list[object] | None = None) -> Path:
    if source.is_file() and not zipfile.is_zipfile(source):
        return source
    holder = temp_dirs if temp_dirs is not None else []
    if source.is_file() and zipfile.is_zipfile(source):
        source = resolve_export_root(load_export_root(source, holder, prefix="rubrics_extract_"))[0]
    candidate = source / "rubrics_d2l.xml"
    if candidate.exists():
        return candidate
    matches = sorted(source.rglob("rubrics_d2l.xml"))
    if len(matches) == 1:
        return matches[0]
    if not matches:
        raise FileNotFoundError("Could not find rubrics_d2l.xml beneath the supplied path.")
    raise FileExistsError("Found multiple rubrics_d2l.xml files; point directly to the desired one.")


def clean_xml_text(value: str) -> str:
    value = unescape(value or "")
    if "<" in value and ">" in value:
        parser = HtmlToTextParser()
        parser.feed(value)
        value = parser.get_text()
    return value.strip()


def cell_text(parent: ET.Element, path: str) -> str:
    elem = parent.find(path)
    if elem is None or elem.text is None:
        return ""
    return clean_xml_text(elem.text)


def score_band_map(rubric: ET.Element) -> dict[str, str]:
    result: dict[str, str] = {}
    for overall in rubric.findall("./overall_level_set/overall_levels/overall_level"):
        name = overall.attrib.get("name", "").strip()
        raw = overall.attrib.get("range_start_value", "").strip()
        if not name:
            continue
        if not raw:
            result[name] = ""
            continue
        number = float(raw)
        rendered = str(int(number)) if number.is_integer() else raw
        suffix = "%" if name == "Not Demonstrated" else "%+"
        result[name] = f"{rendered}{suffix}"
    return result


def parse_levels(rubric: ET.Element) -> list[Level]:
    bands = score_band_map(rubric)
    levels = []
    for level in rubric.findall("./criteria_groups/criteria_group/level_set/levels/level"):
        levels.append(
            Level(
                level_id=level.attrib.get("level_id", ""),
                name=level.attrib.get("name", "").strip(),
                sort_order=int(level.attrib.get("sort_order", "0") or 0),
                score_band=bands.get(level.attrib.get("name", "").strip(), ""),
            )
        )
    return sorted(levels, key=lambda item: item.sort_order)


def set_rubrics_sheet_widths(sheet) -> None:
    widths = {
        "A": 34,
        "B": 42,
        "C": 42,
        "D": 42,
        "E": 42,
        "F": 24,
        "G": 12,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def unique_sheet_title(workbook: Workbook, rubric_name: str) -> str:
    cleaned = "".join(char for char in rubric_name if char not in "[]:*?/\\").strip() or "Rubric"
    base = cleaned[:31]
    if base not in workbook.sheetnames:
        return base

    counter = 2
    while True:
        suffix = f"_{counter}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        if candidate not in workbook.sheetnames:
            return candidate
        counter += 1


def write_rubrics_sheet(sheet, rubric: ET.Element, start_row: int) -> int:
    rubric_name = rubric.attrib.get("name", "").strip() or "Untitled Rubric"
    levels = parse_levels(rubric)

    sheet.cell(start_row, 1, rubric_name)
    sheet.cell(start_row, 1).font = Font(bold=True)
    sheet.cell(start_row, 1).fill = GREEN_FILL
    sheet.cell(start_row, 1).alignment = WRAP_ALIGNMENT

    header_row = start_row + 2
    headers = ["Criteria"]
    for level in levels:
        label = level.name if not level.score_band else f"{level.name}\n{level.score_band}"
        headers.append(label)
    headers.append("Points")

    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, column_index, header)
        cell.alignment = WRAP_ALIGNMENT
        if column_index <= len(levels) + 1:
            cell.fill = GREEN_FILL

    row_index = header_row + 1
    criteria = sorted(
        rubric.findall("./criteria_groups/criteria_group/criteria/criterion"),
        key=lambda elem: int(elem.attrib.get("sort_order", "0") or 0),
    )

    for criterion in criteria:
        sheet.cell(row_index, 1, criterion.attrib.get("name", "").strip())
        sheet.cell(row_index, 1).fill = GREEN_FILL
        sheet.cell(row_index, 1).alignment = WRAP_ALIGNMENT

        cell_map = {
            cell.attrib.get("level_id", ""): cell
            for cell in criterion.findall("./cells/cell")
        }

        max_points: str | float | int = ""
        for column_offset, level in enumerate(levels, start=2):
            rubric_cell = cell_map.get(level.level_id)
            description = ""
            if rubric_cell is not None:
                description = cell_text(rubric_cell, "./description/text")
                points_raw = rubric_cell.attrib.get("cell_value", "").strip()
                if max_points == "" and points_raw:
                    number = float(points_raw)
                    max_points = int(number) if number.is_integer() else points_raw
            cell = sheet.cell(row_index, column_offset, description or None)
            cell.fill = GREEN_FILL
            cell.alignment = WRAP_ALIGNMENT

        points_cell = sheet.cell(row_index, len(levels) + 2, max_points)
        points_cell.alignment = WRAP_ALIGNMENT
        row_index += 1

    return row_index + 1


def write_overall_levels_sheet(sheet, rubrics: list[ET.Element]) -> None:
    headers = ["Rubric", "Overall Level", "Sort Order", "Range Start Value", "Description", "Feedback"]
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(1, column_index, header)
        cell.fill = GREEN_FILL
        cell.font = Font(bold=True)
        cell.alignment = WRAP_ALIGNMENT

    row_index = 2
    for rubric in rubrics:
        rubric_name = rubric.attrib.get("name", "").strip() or "Untitled Rubric"
        overall_levels = sorted(
            rubric.findall("./overall_level_set/overall_levels/overall_level"),
            key=lambda elem: int(elem.attrib.get("sort_order", "0") or 0),
        )
        for overall in overall_levels:
            raw = overall.attrib.get("range_start_value", "").strip()
            rendered = ""
            if raw:
                number = float(raw)
                rendered = int(number) if number.is_integer() else raw
            values = [
                rubric_name,
                overall.attrib.get("name", "").strip(),
                int(overall.attrib.get("sort_order", "0") or 0),
                rendered,
                cell_text(overall, "./description/text"),
                cell_text(overall, "./feedback/text"),
            ]
            for column_index, value in enumerate(values, start=1):
                cell = sheet.cell(row_index, column_index, value if value != "" else None)
                cell.alignment = WRAP_ALIGNMENT
            row_index += 1

    widths = {
        "A": 34,
        "B": 22,
        "C": 12,
        "D": 18,
        "E": 54,
        "F": 42,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _float_or_none(raw: str) -> float | None:
    raw = (raw or "").strip()
    if not raw:
        return None
    try:
        return float(raw)
    except ValueError:
        return None


def rubrics_to_records(xml_path: Path) -> dict:
    """Serialize the rubric grids the workbook path already parses as canonical JSON.

    Extraction mode: attribute values and authored cell/feedback wording are
    preserved verbatim (points also carried as points_raw when non-numeric);
    nothing is inferred. Multiple criteria groups are flattened in document
    order with a diagnostic, matching the workbook rendering.
    """
    root = ET.parse(xml_path).getroot()
    rubric_elems = root.findall("./rubric")
    diagnostics: list[str] = []
    rubrics: list[dict] = []
    for rubric in rubric_elems:
        rubric_name = rubric.attrib.get("name", "").strip() or "Untitled Rubric"
        levels = parse_levels(rubric)
        level_names = {level.level_id: level.name for level in levels}
        groups = rubric.findall("./criteria_groups/criteria_group")
        if len(groups) > 1:
            diagnostics.append(
                f"rubric {rubric_name!r}: {len(groups)} criteria groups flattened in document order"
            )

        criteria: list[dict] = []
        for criterion in sorted(
            rubric.findall("./criteria_groups/criteria_group/criteria/criterion"),
            key=lambda elem: int(elem.attrib.get("sort_order", "0") or 0),
        ):
            cells: list[dict] = []
            for cell in criterion.findall("./cells/cell"):
                level_id = cell.attrib.get("level_id", "")
                points_raw = cell.attrib.get("cell_value", "").strip()
                cells.append(
                    {
                        "level_id": level_id,
                        "level_name": level_names.get(level_id, ""),
                        "points": _float_or_none(points_raw),
                        "points_raw": points_raw,
                        "description": cell_text(cell, "./description/text"),
                    }
                )
            criteria.append(
                {
                    "name": criterion.attrib.get("name", "").strip(),
                    "sort_order": int(criterion.attrib.get("sort_order", "0") or 0),
                    "cells": cells,
                }
            )

        overall_levels: list[dict] = []
        for overall in sorted(
            rubric.findall("./overall_level_set/overall_levels/overall_level"),
            key=lambda elem: int(elem.attrib.get("sort_order", "0") or 0),
        ):
            overall_levels.append(
                {
                    "name": overall.attrib.get("name", "").strip(),
                    "sort_order": int(overall.attrib.get("sort_order", "0") or 0),
                    "range_start_value": _float_or_none(overall.attrib.get("range_start_value", "")),
                    "description": cell_text(overall, "./description/text"),
                    "feedback": cell_text(overall, "./feedback/text"),
                }
            )

        rubrics.append(
            {
                "id": rubric.attrib.get("id", ""),
                "resource_code": rubric.attrib.get("resource_code", ""),
                "name": rubric_name,
                "scoring_method": rubric.attrib.get("scoring_method", ""),
                "attributes": dict(rubric.attrib),
                "description": cell_text(rubric, "./description/text"),
                "levels": [
                    {
                        "level_id": level.level_id,
                        "name": level.name,
                        "sort_order": level.sort_order,
                        "score_band": level.score_band,
                    }
                    for level in levels
                ],
                "criteria": criteria,
                "overall_levels": overall_levels,
            }
        )
    return {
        "schema": RUBRICS_SCHEMA,
        "source_file": xml_path.name,
        "rubrics": rubrics,
        "diagnostics": diagnostics,
    }


def write_rubrics_json(xml_path: Path, json_path: Path) -> dict:
    records = rubrics_to_records(xml_path)
    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(json.dumps(records, indent=2) + "\n", encoding="utf-8")
    return records


def build_workbook(xml_path: Path, output_path: Path, one_sheet_per_rubric: bool = False) -> None:
    root = ET.parse(xml_path).getroot()
    rubrics = root.findall("./rubric")
    if not rubrics:
        raise ValueError("No <rubric> elements found in rubrics_d2l.xml.")

    workbook = Workbook()
    if one_sheet_per_rubric:
        first_sheet = workbook.active
        first_rubric = rubrics[0]
        first_sheet.title = unique_sheet_title(workbook, first_rubric.attrib.get("name", "").strip())
        write_rubrics_sheet(first_sheet, first_rubric, 1)
        set_rubrics_sheet_widths(first_sheet)

        for rubric in rubrics[1:]:
            sheet = workbook.create_sheet(unique_sheet_title(workbook, rubric.attrib.get("name", "").strip()))
            write_rubrics_sheet(sheet, rubric, 1)
            set_rubrics_sheet_widths(sheet)
    else:
        rubrics_sheet = workbook.active
        rubrics_sheet.title = "Rubrics"
        current_row = 1
        for rubric in rubrics:
            current_row = write_rubrics_sheet(rubrics_sheet, rubric, current_row)
        set_rubrics_sheet_widths(rubrics_sheet)

    overall_sheet = workbook.create_sheet("Overall Levels")
    write_overall_levels_sheet(overall_sheet, rubrics)
    workbook.save(output_path)


def main() -> int:
    args = parse_args()
    source = Path(args.source).expanduser().resolve()
    temp_dirs: list[object] = []
    try:
        xml_path = locate_rubrics_xml(source, temp_dirs)
    except Exception as exc:
        print(str(exc), file=sys.stderr)
        return 2

    if args.output:
        output_path = Path(args.output).expanduser().resolve()
    else:
        output_path = xml_path.parent / "Rubrics Extract.xlsx"

    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        build_workbook(xml_path, output_path, one_sheet_per_rubric=args.one_sheet_per_rubric)
    except Exception as exc:
        print(str(exc), file=sys.stderr)
        return 2

    print(output_path)

    if args.json or args.json_output:
        json_path = (
            Path(args.json_output).expanduser().resolve()
            if args.json_output
            else output_path.with_suffix(".json")
        )
        try:
            write_rubrics_json(xml_path, json_path)
        except Exception as exc:
            print(str(exc), file=sys.stderr)
            return 2
        print(json_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
